from django.contrib.auth import authenticate, login
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_protect
from django.db import connection, transaction
from django.shortcuts import render, redirect, get_object_or_404
from django.views.decorators.csrf import csrf_exempt
from django.core.exceptions import ValidationError
from django.contrib import messages
from django.db.models import Q
from django.http import JsonResponse, HttpResponse
from .models import Modulo, DiaSemana, Asignatura, Sala, Profesor, Horario, Licencia, Reemplazos,Recuperacion, Carrera
from django.core.exceptions import ObjectDoesNotExist
import json
import openpyxl
from openpyxl.styles import NamedStyle, Border, Alignment
import logging
import locale
import os
from datetime import timedelta, datetime
from collections import defaultdict

locale.setlocale(locale.LC_TIME, 'es_ES.UTF-8')

def login_view(request):
    if request.method == 'POST':
        username = request.POST['usuario']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('reemplazos')
        else:
            messages.error(request, 'Usuario o contraseña incorrectos.')
    return render(request, 'templates/login.html')

def docente_view(request):
    profesores = Profesor.objects.all()
    context = {
        'profesores': profesores,
    }

    if request.method == 'POST' and 'id_profesor' in request.POST:
        id_profesor = request.POST.get('id_profesor')
        if id_profesor:
            try:
                profesor_a_eliminar = get_object_or_404(Profesor, id_profesor=id_profesor)
                horarios_asociados = Horario.objects.filter(profesor_id_profesor=profesor_a_eliminar)
                
                for horario in horarios_asociados:
                    Reemplazos.objects.filter(horario=horario).delete()
                    
                Licencia.objects.filter(profesor_id_profesor=profesor_a_eliminar).delete()
                profesor_a_eliminar.delete()

                messages.success(request, "Profesor y sus licencias eliminadas exitosamente.")
                return redirect('docente')

            except Profesor.DoesNotExist:
                messages.error(request, "El profesor no existe.")
                return redirect('docente')

    return render(request, 'templates/docente.html', context)

def guardar_licencia(request):
    if request.method == 'POST':
        profesor_id = request.POST.get('profesor_id_profesor')  
        fecha_inicio = request.POST.get('fecha_inicio')
        fecha_termino = request.POST.get('fecha_fin')
        motivo = request.POST.get('motivo')
        observaciones = request.POST.get('observaciones')

        if not profesor_id:
            return JsonResponse({'success': False, 'error': 'Falta el ID del profesor'})
        
        profesor = get_object_or_404(Profesor, id_profesor=profesor_id)

        try:
            nueva_licencia = Licencia.objects.create(
                profesor_id_profesor=profesor,
                fecha_inicio=fecha_inicio,
                fecha_termino=fecha_termino,
                motivo=motivo,
                observaciones=observaciones,
                estado="sin_asignar"
            )
            return JsonResponse({'success': True, 'message': 'Licencia guardada exitosamente.'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': f'Ocurrió un error: {str(e)}'})

    return JsonResponse({'success': False, 'error': 'Método no permitido'})

def recuperacion_view(request):
    recuperaciones = Recuperacion.objects.all()

    return render(request, 'templates/gestion_recuperacion.html', {
        'recuperaciones': recuperaciones
    })        

def eliminar_recuperacion(request, id_recuperacion):
    try:
        recuperacion = Recuperacion.objects.get(id_recuperacion=id_recuperacion)
        recuperacion.delete()
        return JsonResponse({'success': True, 'message': 'Recuperación eliminada correctamente'})
    except Recuperacion.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Recuperación no encontrada'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)
    
@csrf_protect  
def actualizar_recuperacion(request, id_recuperacion):
    if request.method == 'PUT': 
        try:
            data = json.loads(request.body)
            
            if not data.get('numero_modulos') or not data.get('fecha_clase') or not data.get('fecha_recuperacion') or not data.get('hora_recuperacion') or not data.get('sala'):
                return JsonResponse({'success': False, 'message': 'Faltan datos requeridos.'})
            
            recuperacion = Recuperacion.objects.get(id=id_recuperacion)
            recuperacion.numero_modulos = data.get('numero_modulos')
            recuperacion.fecha_clase = data.get('fecha_clase')
            recuperacion.fecha_recuperacion = data.get('fecha_recuperacion')
            recuperacion.hora_recuperacion = data.get('hora_recuperacion')
            recuperacion.sala = data.get('sala')
            recuperacion.save()

            return JsonResponse({'success': True, 'message': 'Recuperación actualizada correctamente.'})

        except Recuperacion.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Recuperación no encontrada.'})

        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'message': 'Error al procesar los datos JSON.'})

        except Exception as e:
            return JsonResponse({'success': False, 'message': f'Error inesperado: {str(e)}'})
    
    else:
        return JsonResponse({'success': False, 'message': 'Método no permitido. Solo se permite PUT.'})

def reportes_view(request):
    profesores = Profesor.objects.all()

    profesores_concatenados = []

    for profesor in profesores:
        nombre_completo = f"{profesor.nombre} {profesor.segundo_nombre} {profesor.apellido} {profesor.segundo_apellido}"
        profesores_concatenados.append({
            'id': profesor.id_profesor,
            'nombre_completo': nombre_completo
        })

    context = {
        'profesores': profesores_concatenados,
    }

    return render(request, 'templates/reportes.html', context)


@login_required(login_url='/login/') 
def base_view(request):
    return render(request, 'base.html')

def CustomLogoutView(request):
    return render(request, 'templates/login.html')

def asignatura_view(request):
    if request.method == 'GET':
        asignaturas = Asignatura.objects.all().order_by('nombre_asignatura').values('id_asignatura', 'nombre_asignatura')
        asignaturas_list = list(asignaturas)
        return JsonResponse(asignaturas_list, safe=False)

    elif request.method == 'POST':
        data = json.loads(request.body)
        nombre_asignatura = data.get('nombre')
        
        if nombre_asignatura:
            nueva_asignatura = Asignatura(nombre_asignatura=nombre_asignatura)
            nueva_asignatura.save()
            return JsonResponse({
                'message': 'Asignatura agregada',
                'nombre': nueva_asignatura.nombre_asignatura,
                'id': nueva_asignatura.id_asignatura}, status=201)          
        else:
            return JsonResponse({'error': 'Nombre de asignatura no proporcionado'}, status=400)

    return JsonResponse({'error': 'Método no permitido'}, status=405)

def sala_view(request):
    if request.method == 'GET':
        salas = Sala.objects.all().order_by('numero_sala').values('id_sala', 'numero_sala')
        salas_list = list(salas)
        return JsonResponse(salas_list, safe=False)
    
    elif request.method == 'POST':
        data = json.loads(request.body)
        numero_sala = data.get('numero_sala')
        if not numero_sala:
            return JsonResponse({'message': 'Número de la sala no proporcionado'}, status=400)
        
        nueva_sala = Sala(numero_sala=numero_sala)
        nueva_sala.save()
        return JsonResponse({
            'message': 'Sala agregada',
            'id_sala': nueva_sala.id_sala,
            'numero_sala': nueva_sala.numero_sala
        }, status=201)
    return JsonResponse({'error': 'Método no permitido'}, status=405)

def carrera_view(request):
    if request.method == 'GET':
        carreras = Carrera.objects.all().order_by('nombre_carrera').values('id_carrera', 'nombre_carrera')
        carreras_list = list(carreras)
        return JsonResponse(carreras_list, safe=False)

    elif request.method == 'POST':
        data = json.loads(request.body)
        nombre_carrera = data.get('nombre')
        
        if nombre_carrera:
            nueva_carrera = Carrera(nombre_carrera=nombre_carrera)
            nueva_carrera.save()
            return JsonResponse({
                'message': 'Carrera agregada',
                'nombre': nueva_carrera.nombre_carrera,
                'id': nueva_carrera.id_carrera}, status=201)          
        else:
            return JsonResponse({'error': 'Nombre de carrera no proporcionado'}, status=400)

    return JsonResponse({'error': 'Método no permitido'}, status=405)

def crear_profesor_y_horarios(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            profesor_data = data.get('profesor')

            if not profesor_data.get('nombre') or not profesor_data.get('apellido'):
                logger.warning("Faltan datos del profesor: nombre o apellido no proporcionados.")
                return JsonResponse({"error": "Faltan datos del profesor"}, status=400)

            ultimo_profesor = Profesor.objects.last()
            nuevo_id_profesor = ultimo_profesor.id_profesor + 1 if ultimo_profesor else 1
            logger.info(f"Creando nuevo profesor con ID: {nuevo_id_profesor}")

            with transaction.atomic():
                profesor = Profesor.objects.create(
                    id_profesor=nuevo_id_profesor,
                    nombre=profesor_data['nombre'],
                    apellido=profesor_data['apellido'],
                    segundo_nombre=profesor_data.get('segundo_nombre', ''),
                    segundo_apellido=profesor_data.get('segundo_apellido', '')
                )
                logger.info(f"Profesor {profesor.nombre} {profesor.apellido} creado exitosamente.")

                for horario_data in data['horarios_asignados']:
                    try:
                        asignatura = Asignatura.objects.get(id_asignatura=horario_data['asignaturaID'])
                        sala = Sala.objects.get(id_sala=horario_data['salaID'])
                        modulo = Modulo.objects.get(id_modulo=horario_data['moduloID'])
                        dia = DiaSemana.objects.get(id_dia=horario_data['diaID'])
                        carrera = Carrera.objects.get(id_carrera=horario_data['carreraID'])
                    except Asignatura.DoesNotExist:
                        logger.error(f"Asignatura con ID {horario_data['asignaturaID']} no existe.")
                        return JsonResponse({"error": f"Asignatura con ID {horario_data['asignaturaID']} no existe."}, status=400)
                    except Sala.DoesNotExist:
                        logger.error(f"Sala con ID {horario_data['salaID']} no existe.")
                        return JsonResponse({"error": f"Sala con ID {horario_data['salaID']} no existe."}, status=400)
                    except Modulo.DoesNotExist:
                        logger.error(f"Módulo con ID {horario_data['moduloID']} no existe.")
                        return JsonResponse({"error": f"Módulo con ID {horario_data['moduloID']} no existe."}, status=400)
                    except DiaSemana.DoesNotExist:
                        logger.error(f"Día de semana con ID {horario_data['diaID']} no existe.")
                        return JsonResponse({"error": f"Día de semana con ID {horario_data['diaID']} no existe."}, status=400)
                    except Carrera.DoesNotExist:
                        logger.error(f"Carrera con ID {horario_data['carreraID']} no existe.")
                        return JsonResponse({"error": f"Carrera con ID {horario_data['carreraID']} no existe."}, status=400)

                    Horario.objects.create(
                        profesor_id_profesor=profesor,
                        asignatura_id_asignatura=asignatura,
                        sala_id_sala=sala,
                        dia_semana_id_dia=dia,
                        modulo_id_modulo=modulo,
                        seccion=horario_data['seccion'],
                        jornada=horario_data['jornada'],
                        carrera_id_carrera=carrera
                    )
                    logger.info(f"Horario creado para el profesor {profesor.nombre} {profesor.apellido}, seccion {horario_data['seccion']}.")

                return JsonResponse({"message": "Profesor y horarios creados exitosamente!"}, status=201)

        except Exception as e:
            logger.error(f"Error al crear profesor y horarios: {str(e)}")
            return JsonResponse({"error": str(e)}, status=500)

    return JsonResponse({'error': 'Método no permitido'}, status=405)

def editar_profesor(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            profesor_data = data.get('profesor')
            if not profesor_data.get('id') or not profesor_data.get('nombre') or not profesor_data.get('apellido'):
                return JsonResponse({"error": "Faltan datos del profesor"}, status=400)

            try:
                profesor = Profesor.objects.get(id_profesor=profesor_data['id'])
            except Profesor.DoesNotExist:
                return JsonResponse({"error": f"Profesor con ID {profesor_data['id']} no existe."}, status=400)

            profesor.nombre = profesor_data['nombre']
            profesor.apellido = profesor_data['apellido']
            profesor.segundo_nombre = profesor_data.get('segundo_nombre', '')
            profesor.segundo_apellido = profesor_data.get('segundo_apellido', '')
            profesor.save()

            logger.info(f"Profesor {profesor.nombre} {profesor.apellido} actualizado exitosamente.")

            horarios_actuales = Horario.objects.filter(profesor_id_profesor=profesor)

            horarios_enviados = [
                (horario_data['asignaturaID'], horario_data['salaID'], horario_data['moduloID'], horario_data['diaID'], horario_data['seccion'], horario_data['carreraID'])
                for horario_data in data['horarios_asignados']
            ]

            for horario_actual in horarios_actuales:
                horario_key = (horario_actual.asignatura_id_asignatura.id_asignatura, 
                               horario_actual.sala_id_sala.id_sala, 
                               horario_actual.modulo_id_modulo.id_modulo, 
                               horario_actual.dia_semana_id_dia.id_dia, 
                               horario_actual.seccion,
                               horario_actual.carrera_id_carrera.id_carrera)

                if horario_key not in horarios_enviados:
                    horario_actual.activo = 0
                    horario_actual.save()
                    logger.info(f"Desactivando horario para el profesor {profesor.nombre} {profesor.apellido}: "
                                f"Asignatura ID {horario_actual.asignatura_id_asignatura.id_asignatura}, "
                                f"Sala ID {horario_actual.sala_id_sala.id_sala}, "
                                f"Módulo ID {horario_actual.modulo_id_modulo.id_modulo}, "
                                f"Día ID {horario_actual.dia_semana_id_dia.id_dia}, "
                                f"Sección {horario_actual.seccion}, "
                                f"Carrera ID {horario_actual.carrera_id_carrera.id_carrera}")


            for horario_data in data['horarios_asignados']:
                try:
                    asignatura = Asignatura.objects.get(id_asignatura=horario_data['asignaturaID'])
                    sala = Sala.objects.get(id_sala=horario_data['salaID'])
                    modulo = Modulo.objects.get(id_modulo=horario_data['moduloID'])
                    dia = DiaSemana.objects.get(id_dia=horario_data['diaID'])
                    carrera = Carrera.objects.get(id_carrera=horario_data['carreraID'])
                except Asignatura.DoesNotExist:
                    return JsonResponse({"error": f"Asignatura con ID {horario_data['asignaturaID']} no existe."}, status=400)
                except Sala.DoesNotExist:
                    return JsonResponse({"error": f"Sala con ID {horario_data['salaID']} no existe."}, status=400)
                except Modulo.DoesNotExist:
                    return JsonResponse({"error": f"Módulo con ID {horario_data['moduloID']} no existe."}, status=400)
                except DiaSemana.DoesNotExist:
                    return JsonResponse({"error": f"Día de semana con ID {horario_data['diaID']} no existe."}, status=400)
                except Carrera.DoesNotExist:
                    return JsonResponse({"error": f"Carrera con ID {horario_data['carreraID']} no existe."}, status=400)

                try:
                    horario_existente = Horario.objects.get(
                        profesor_id_profesor=profesor,
                        asignatura_id_asignatura=asignatura,
                        sala_id_sala=sala,
                        dia_semana_id_dia=dia,
                        modulo_id_modulo=modulo,
                        carrera_id_carrera=carrera,
                        seccion=horario_data['seccion']
                    )
                    if horario_existente.jornada != horario_data['jornada']:
                        horario_existente.jornada = horario_data['jornada']
                        horario_existente.save()
                        logger.info(f"Jornada del horario para el profesor {profesor.nombre} {profesor.apellido} "
                                    f"actualizada a {horario_data['jornada']}.")
                except Horario.DoesNotExist:
                    Horario.objects.create(
                        profesor_id_profesor=profesor,
                        asignatura_id_asignatura=asignatura,
                        sala_id_sala=sala,
                        dia_semana_id_dia=dia,
                        modulo_id_modulo=modulo,
                        carrera_id_carrera=carrera,
                        seccion=horario_data['seccion'],
                        jornada=horario_data['jornada']
                    )
                    logger.info(f"Horario creado para el profesor {profesor.nombre} {profesor.apellido}: "
                                f"Asignatura ID {asignatura.id_asignatura}, Sala ID {sala.id_sala}, "
                                f"Módulo ID {modulo.id_modulo}, Día ID {dia.id_dia}, "
                                f"Sección {horario_data['seccion']}, Jornada {horario_data['jornada']}, "
                                f"Carrera ID {carrera.id_carrera}.")
            
            return JsonResponse({"message": "Profesor y horarios actualizados exitosamente!"}, status=200)

        except Exception as e:
            logger.error(f"Error al actualizar el profesor: {str(e)}")
            return JsonResponse({"error": str(e)}, status=500)
   
def crear_docente_view(request):
    modulos = Modulo.objects.all()
    dias = DiaSemana.objects.all()
    profesores = Profesor.objects.all()
    ultimo_profesor = profesores.last()
    ultimo_id = ultimo_profesor.id_profesor if ultimo_profesor else 0
    context = {
        'modulos': modulos,
        'dias': dias,
        'profesores': profesores,
        'ultimo_id': ultimo_id
    }
    if request.method == 'POST':
        if 'eliminar' in request.POST:
            id_profesor = request.POST.get('id_profesor')
            try:
                profesor_a_eliminar = get_object_or_404(Profesor, id_profesor=id_profesor)
                profesor_a_eliminar.delete()
                messages.success(request, "Profesor eliminado exitosamente.")
            except Profesor.DoesNotExist:
                messages.error(request, "El profesor no existe.")
        else:
            nombre = request.POST.get('primer_nombre').strip().capitalize()
            segundo_nombre = request.POST.get('segundo_nombre').strip().capitalize() 
            apellido = request.POST.get('primer_apellido').strip().capitalize()
            segundo_apellido = request.POST.get('segundo_apellido').strip().capitalize()

            if not nombre or not apellido:
                messages.error(request, 'Los nombres y apellidos son obligatorios.')
                return redirect('docente')

            nuevo_profesor = Profesor(
                nombre=nombre,
                segundo_nombre=segundo_nombre,
                apellido=apellido,
                segundo_apellido=segundo_apellido
            )
            nuevo_profesor.save()
            messages.success(request, "Profesor agregado exitosamente.")
    return render(request, 'templates/crear_docente.html',context)



def registrar_recuperacion(request):
    if request.method == 'POST':
        try:
            logger.info("Iniciando registro de recuperación.")
            data = json.loads(request.body)
            logger.debug(f"Datos recibidos: {data}")
            profesor = data.get('profesorNombre')
            asignatura = data.get('asignaturaNombre')
            numero_modulos = data.get('numero_modulos')
            fecha_clase = data.get('fecha_clase')
            fecha_recuperacion = data.get('fecha_recuperacion')
            hora_recuperacion = data.get('hora_recuperacion')
            sala = data.get('salaNombre')
            if not all([profesor, asignatura, numero_modulos, fecha_clase, fecha_recuperacion, hora_recuperacion, sala]):
                logger.warning("Faltan datos en el formulario.")
                return JsonResponse({'success': False, 'message': 'Faltan datos en el formulario.'}, status=400)
            recuperacion = Recuperacion(
                profesor=profesor,
                asignatura=asignatura,
                numero_modulos=numero_modulos,
                fecha_clase=fecha_clase,
                fecha_recuperacion=fecha_recuperacion,
                hora_recuperacion=hora_recuperacion,
                sala=sala
            )
            recuperacion.save()
            return JsonResponse({'success': True, 'message': 'Recuperación registrada correctamente.'})

        except Exception as e:
            logger.error(f"Ocurrió un error al registrar la recuperación: {str(e)}", exc_info=True)
            return JsonResponse({'success': False, 'message': f'Ocurrió un error: {str(e)}'}, status=500)
    
    logger.warning("Método no permitido para la solicitud.")
    return JsonResponse({'success': False, 'message': 'Método no permitido'}, status=405)

def reemplazos_view(request):
    sql_query = '''
    select r.semana, 
    concat(p.nombre,' ',IF(p.segundo_nombre IS NOT NULL, CONCAT(p.segundo_nombre, ' '),''), p.apellido, IF(p.segundo_apellido IS NOT NULL, CONCAT(' ',p.segundo_apellido),'')) as profesor,
    a.nombre_asignatura,
    h.seccion,
    s.numero_sala,
        CONCAT(
            SUBSTRING_INDEX(GROUP_CONCAT(mo.hora_modulo ORDER BY mo.hora_modulo SEPARATOR ', '), '-', 1), 
            '-',
            SUBSTRING_INDEX(SUBSTRING_INDEX(GROUP_CONCAT(mo.hora_modulo ORDER BY mo.hora_modulo SEPARATOR ', '), ',', -1), '-', -1)
        ) AS modulo,
    count(*) as registros,
    ds.nombre_dia,
    r.fecha_reemplazo, 
    r.profesor_reemplazo,
    GROUP_CONCAT(r.id_reemplazo ORDER BY r.id_reemplazo SEPARATOR '-') as id_reemplazos,
 	IF(COUNT(mo.id_modulo) > 1, GROUP_CONCAT(mo.id_modulo ORDER BY mo.id_modulo SEPARATOR '-'), MAX(mo.id_modulo)) AS id_modulo,
    ds.id_dia
    from test.reemplazos r
    join test.horario h on h.id_horario = r.horario_id_horario 
    join test.profesor p on p.id_profesor = h.profesor_id_profesor
    join test.sala s on s.id_sala = h.sala_id_sala 
    join test.modulo mo on mo.id_modulo = h.modulo_id_modulo
    join test.asignatura a on a.id_asignatura = h.asignatura_id_asignatura 
    join test.dia_semana ds on ds.id_dia = h.dia_semana_id_dia 
    group by r.semana, r.fecha_reemplazo, r.profesor_reemplazo, s.numero_sala, h.seccion, h.profesor_id_profesor, h.dia_semana_id_dia 
    order by r.fecha_reemplazo 
    '''

    with connection.cursor() as cursor:
        cursor.execute(sql_query)
        rows = cursor.fetchall()

    reemplazos_listados = []

    for row in rows:
        semana = row[0]
        profesor_nombre = row[1] 
        nombre_asignatura = row[2] 
        seccion = row[3]
        numero_sala = row[4]
        hora_modulo = row[5]
        registros = row[6]
        nombre_dia = row[7]
        fecha_reemplazo = row[8]
        profesor_reemplazo = row[9]
        id_reemplazo = row[10]
        id_modulo= row[11]
        id_dia= row[12]

        reemplazos_listados.append({
            'semana': semana,
            'profesor_nombre': profesor_nombre,
            'nombre_asignatura': nombre_asignatura,
            'seccion': seccion,
            'numero_sala': numero_sala,
            'hora_modulo': hora_modulo,
            'registros': registros,
            'nombre_dia': nombre_dia,
            'fecha_reemplazo': fecha_reemplazo,
            'profesor_reemplazo': profesor_reemplazo,
            'id_reemplazo': id_reemplazo,
            'id_modulo': id_modulo,
            'id_dia':id_dia
        })

    return render(request, 'templates/gestion_reemplazo.html', {
        'reemplazos': reemplazos_listados
    })

def profesores_con_licencia_no_asignada(request):
    profesores = Profesor.objects.filter(
        licencia__estado='sin_asignar'
    ).distinct()
    profesores_data = []
    for profesor in profesores:
        licencia = Licencia.objects.filter(profesor_id_profesor=profesor, estado='sin_asignar').first()
        if licencia:
            profesores_data.append({
                'id_profesor': profesor.id_profesor,
                'nombre': profesor.nombre,
                'segundo_nombre': profesor.segundo_nombre,
                'apellido': profesor.apellido,
                'segundo_apellido': profesor.segundo_apellido,
                'id_licencia': licencia.id_licencia,
                'fecha_inicio': licencia.fecha_inicio.strftime('%Y-%m-%d'),
                'fecha_termino': licencia.fecha_termino.strftime('%Y-%m-%d')
            })

    return JsonResponse({'profesores': profesores_data}, safe=False)


logger = logging.getLogger(__name__)
def obtener_clases_por_docente(request):
    docente_id = request.GET.get('docente_id')
    fecha_inicio = request.GET.get('fecha_inicio')  
    fecha_termino = request.GET.get('fecha_termino')

    if not docente_id or not fecha_inicio or not fecha_termino:
        logger.error("Faltan parámetros en la solicitud.")
        return JsonResponse({'error': 'Faltan parámetros'}, status=400)

    try:
        fecha_inicio = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
        fecha_termino = datetime.strptime(fecha_termino, '%Y-%m-%d').date()
        if fecha_termino < fecha_inicio:
            logger.warning("La fecha de término no puede ser antes de la fecha de inicio.")
            return JsonResponse({'error': 'La fecha de término no puede ser antes de la fecha de inicio'}, status=400)
        try:
            profesor = Profesor.objects.get(id_profesor=docente_id)
            logger.debug(f"Profesor encontrado: {profesor.nombre} {profesor.apellido}")
        except Profesor.DoesNotExist:
            logger.error(f"Profesor con id {docente_id} no encontrado.")
            return JsonResponse({'error': 'Profesor no encontrado'}, status=404)
        
        horarios = Horario.objects.filter(
            profesor_id_profesor=profesor, activo=True
        ).filter(Q(dia_semana_id_dia__gte=1) & Q(dia_semana_id_dia__lte=6))

        logger.debug(f"Horarios obtenidos: {horarios.count()} horarios")

        clases = []
        current_date = fecha_inicio
        while current_date <= fecha_termino:
            dia_semana = current_date.weekday() + 1
            logger.debug(f"Procesando fecha: {current_date} (Día de la semana: {dia_semana})")
            for horario in horarios.filter(dia_semana_id_dia=dia_semana,):
                dia = current_date.strftime('%A').capitalize()
                clases.append({
                    'asignatura': horario.asignatura_id_asignatura.nombre_asignatura,
                    'seccion': horario.seccion,
                    'sala': horario.sala_id_sala.numero_sala,
                    'modulo': horario.modulo_id_modulo.hora_modulo,
                    'dia': dia,
                    'fecha_clase': current_date.strftime('%d-%m-%Y'),
                    'modulo_id': horario.modulo_id_modulo.id_modulo,
                    'dia_semana_id': horario.dia_semana_id_dia.id_dia,
                    'id_horario': horario.id_horario,
                    'profesor_id': horario.profesor_id_profesor.id_profesor
                })
            current_date += timedelta(days=1)

        if not clases:
            logger.info(f"No hay clases para el profesor {profesor.nombre} {profesor.apellido} durante el rango de fechas.")
        
        logger.debug(f"Clases encontradas: {len(clases)}")
        return JsonResponse({'clases': clases})

    except Exception as e:
        logger.error(f"Error inesperado: {str(e)}")
        return JsonResponse({'error': f'Error: {str(e)}'}, status=500)
    
def obtener_profesores_disponibles(request):
    dia_semana = request.GET.get('dia_semana')
    modulo_id = request.GET.get('modulo_id')
    if not dia_semana or not modulo_id:
        return JsonResponse({'error': 'Faltan parámetros'}, status=400)
    try:
        dia_semana = int(dia_semana)
        modulo_id = int(modulo_id)
    except ValueError:
        return JsonResponse({'error': 'Los parámetros deben ser números enteros'}, status=400)
    profesores_disponibles = Profesor.objects.exclude(
        horario__dia_semana_id_dia=dia_semana,
        horario__modulo_id_modulo_id=modulo_id
    ).distinct()
    profesores_data = list(profesores_disponibles.values('id_profesor', 'nombre', 'apellido', 'segundo_nombre', 'segundo_apellido'))
    return JsonResponse({'profesores': profesores_data}, safe=False)

def registrar_reemplazo(request):
    if request.method == 'POST':
        try:
            logger.info('Recibida solicitud POST para registrar reemplazos.')
            data = json.loads(request.body)
            reemplazos = data.get('reemplazos', [])
            id_licencia = data.get('id_licencia')

            if not reemplazos:
                logger.warning('No se proporcionaron reemplazos en la solicitud.')
                return JsonResponse({'error': 'No se proporcionaron reemplazos'}, status=400)

            if not id_licencia:
                logger.warning('No se proporcionó id_licencia en la solicitud.')
                return JsonResponse({'error': 'No se proporcionó id_licencia'}, status=400)
            licencia = get_object_or_404(Licencia, id_licencia=id_licencia)
            logger.info(f'Licencia {id_licencia} actualizada a estado "asignado".')
            logger.info(f'Número de reemplazos recibidos: {len(reemplazos)}')
            for reemplazo in reemplazos:
                semana = reemplazo.get('semana')
                fecha_reemplazo = reemplazo.get('fecha_reemplazo')
                profesor_reemplazo = reemplazo.get('profesor_reemplazo')
                horario_id = reemplazo.get('horario_id')
                logger.info(f'Procesando reemplazo: Semana={semana}, Fecha={fecha_reemplazo}, Profesor={profesor_reemplazo}, Horario ID={horario_id}')
                try:
                    horario = Horario.objects.get(id_horario=horario_id)
                    logger.info(f'Horario encontrado: {horario}')
                    licencia.estado = 'asignado'
                    licencia.save()
                except Horario.DoesNotExist:
                    logger.error(f'Horario no encontrado para ID: {horario_id}')
                    return JsonResponse({'error': f'Horario no encontrado para ID {horario_id}'}, status=400)
                
                reemplazo_obj = Reemplazos.objects.create(
                    semana=semana,
                    fecha_reemplazo=fecha_reemplazo,
                    profesor_reemplazo=profesor_reemplazo,
                    horario=horario,
                )
                logger.info(f'Reemplazo creado: {reemplazo_obj}')

            return JsonResponse({'success': True}, status=200)
        except json.JSONDecodeError:
            logger.error('Error al decodificar el JSON recibido.')
            return JsonResponse({'error': 'Error al procesar la solicitud. JSON mal formado.'}, status=400)
        except Exception as e:
            logger.error(f'Error inesperado: {str(e)}')
            return JsonResponse({'error': f'Error inesperado: {str(e)}'}, status=500)
    
    logger.warning('Método no permitido. Solo se acepta POST.')
    return JsonResponse({'error': 'Método no permitido'}, status=405)


def modificar_docente_view(request,id):
    modulos = Modulo.objects.all()
    dias = DiaSemana.objects.all()
    carrera = Carrera.objects.all()
    horarios = Horario.objects.filter(profesor_id_profesor=id)
    profesor = get_object_or_404(Profesor, id_profesor=id)
    context = {
        'modulos': modulos,
        'dias': dias,
        'profesor':profesor,
        'horario:':horarios,
        'carrera':carrera
    }
    return render(request, 'templates/modificar_docente.html', context)

def modificar_profesor_y_horarios(request):
    if request.method == 'PUT':
        try:
            data = json.loads(request.body)
            logger.info(f"Datos recibidos: {data}")
            profesor_data = data.get('profesor')
            profesor_id = profesor_data.get('id_profesor')

            if not profesor_id:
                return JsonResponse({"error": "ID del profesor no proporcionado."}, status=400)

            profesor = get_object_or_404(Profesor, id_profesor=profesor_id)

            profesor.nombre = profesor_data['nombre']
            profesor.apellido = profesor_data['apellido']
            profesor.segundo_nombre = profesor_data.get('segundo_nombre', '')
            profesor.segundo_apellido = profesor_data.get('segundo_apellido', '')
            profesor.save()

            for horario_data in data.get('horarios_asignados', []):
                logger.info(f"Procesando horario: {horario_data}")

                try:
                    asignatura = Asignatura.objects.get(id_asignatura=horario_data['asignaturaID'])
                    sala = Sala.objects.get(id_sala=horario_data['salaID'])
                    modulo = Modulo.objects.get(id_modulo=horario_data['moduloID'])
                    dia = DiaSemana.objects.get(id_dia=horario_data['diaID'])
                    carrera = Carrera.objects.get(id_carrera=horario_data['carreraID'])

                    if not carrera:
                        logger.error(f"Carrera no encontrada: {horario_data['carreraID']}")
                        return JsonResponse({"error": "Carrera no encontrada."}, status=400)

                except Exception as e:
                    logger.error(f"Error al obtener los objetos relacionados: {str(e)}")
                    return JsonResponse({"error": f"Error al obtener los objetos relacionados: {str(e)}"}, status=500)

                horario, created = Horario.objects.update_or_create(
                    profesor_id_profesor=profesor,
                    asignatura_id_asignatura=asignatura,
                    sala_id_sala=sala,
                    dia_semana_id_dia=dia,
                    modulo_id_modulo=modulo,
                    carrera_id_carrera=carrera,
                    defaults={
                        'seccion': horario_data['seccion'],
                        'jornada': horario_data['jornada']
                    }
                )

            return JsonResponse({"message": "Profesor y horarios actualizados exitosamente!"}, status=200)

        except Exception as e:
            logger.error(f"Error al modificar profesor y horarios: {str(e)}")
            return JsonResponse({"error": str(e)}, status=500)

    return JsonResponse({'error': 'Método no permitido'}, status=405)

def docente_recuperación(request):
    profesores = Profesor.objects.filter(horario__isnull=False).distinct()
    profesores_data = list(profesores.values(
            'id_profesor', 
            'nombre', 
            'apellido', 
            'segundo_nombre', 
            'segundo_apellido'            
        ))
    return JsonResponse({'profesores': profesores_data})

def docente_asignatura(request,profesorId):
    asignaturas = Asignatura.objects.filter(horario__profesor_id_profesor=profesorId).distinct()
    asignatura_data = list(asignaturas.values(
        'id_asignatura',
        'nombre_asignatura'
    ))
    
    return JsonResponse({'asignaturas': asignatura_data })

def todas_asignaturas(request):
    asignaturas = Asignatura.objects.all()  
    asignaturas_data = list(asignaturas.values(
        'id_asignatura',
        'nombre_asignatura'
    ))
    return JsonResponse({'todas_asignaturas': asignaturas_data})


def todas_salas(request):
    salas = Sala.objects.all()  
    salas_data = list(salas.values(
        'id_sala', 
        'numero_sala'
    ))
    return JsonResponse({'todas_salas': salas_data})

def todas_carreras(request):
    carreras = Carreras.objects.all()
    return JsonResponse(carreras)

def licencias_profesores(request):
    licencias = Licencia.objects.select_related('profesor_id_profesor').all()
    licencias_info = [
        {
            'id_licencia': licencia.id_licencia,
            'motivo': licencia.motivo,
            'observaciones': licencia.observaciones,
            'fecha_inicio': licencia.fecha_inicio.strftime('%d/%m'),
            'fecha_termino': licencia.fecha_termino.strftime('%d/%m'),
            'fi': licencia.fecha_inicio,
            'ft': licencia.fecha_termino,
            'estado': licencia.estado,
            'nombre_profesor': f"{licencia.profesor_id_profesor.nombre} {licencia.profesor_id_profesor.segundo_nombre} {licencia.profesor_id_profesor.apellido} {licencia.profesor_id_profesor.segundo_apellido}",
        }
        for licencia in licencias
    ]

    return JsonResponse(licencias_info, safe=False)

def gestionar_licencias(request, profesor_id):
    profesor = get_object_or_404(Profesor, id_profesor=profesor_id)

    licencias = Licencia.objects.filter(profesor_id_profesor=profesor_id)

    return render(request, 'templates/gestionar_licencias.html', {
        'profesor': profesor,
        'licencias': licencias,
    })


def editar_licencia(request, id_licencia):
    if request.method == 'POST':  
        try:
            licencia = Licencia.objects.get(id_licencia=id_licencia)

            licencia.fecha_inicio = request.POST.get('fecha_inicio')
            licencia.fecha_termino = request.POST.get('fecha_termino')
            licencia.motivo = request.POST.get('motivo')
            licencia.observaciones = request.POST.get('observaciones')

            licencia.save()

            return redirect('gestionar_licencias', profesor_id=licencia.profesor_id_profesor.id_profesor)

        except Licencia.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Licencia no encontrada.'})

        except Exception as e:
            return JsonResponse({'success': False, 'message': f'Error inesperado: {str(e)}'})

    elif request.method == 'GET':  
        try:
            licencia = Licencia.objects.get(id_licencia=id_licencia)
            data = {
                'id_licencia': licencia.id_licencia,
                'fecha_inicio': licencia.fecha_inicio,
                'fecha_termino': licencia.fecha_termino,
                'motivo': licencia.motivo,
                'observaciones': licencia.observaciones
            }
            return JsonResponse(data)
        except Licencia.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Licencia no encontrada.'})

    return JsonResponse({'success': False, 'message': 'Método no permitido.'})

    
def eliminar_licencia(request, id_licencia):
    if request.method == 'DELETE':
        try:
            licencia = Licencia.objects.get(id_licencia=id_licencia)
            licencia.delete()
            return JsonResponse({'success': True, 'message': 'Licencia eliminada correctamente'})
        except Licencia.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Licencia no encontrada'}, status=404)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)
    return JsonResponse({'success': False, 'message': 'Método no permitido'}, status=405)


def profesor_por_nombre(request, nombre):
    logger.info(f'Buscando profesor con el nombre: {nombre}')
    nombre_parts = nombre.split()
    logger.info(f'Nombre dividido en partes: {nombre_parts}')
    filters = Q()
    for part in nombre_parts:
        logger.info(f'Buscando la parte: {part}')
        filters |= Q(nombre__icontains=part) | Q(segundo_nombre__icontains=part) | Q(apellido__icontains=part) | Q(segundo_apellido__icontains=part)
    try:
        profesores = Profesor.objects.filter(filters)
        logger.info(f'Profesores encontrados: {profesores.count()}')
        
        if profesores.exists():
            profesor_data = [{
                'id_profesor': p.id_profesor,
                'nombre': p.nombre,
                'segundo_nombre': p.segundo_nombre,
                'apellido': p.apellido,
                'segundo_apellido': p.segundo_apellido,
            } for p in profesores]
            
            return JsonResponse({'profesores': profesor_data}, status=200)
        else:
            logger.warning('No se encontró ningún profesor.')
            return JsonResponse({'mensaje': 'No se encontró el profesor'}, status=404)
    
    except Exception as e:
        logger.error(f'Error al realizar la búsqueda: {str(e)}')
        return JsonResponse({'mensaje': 'Error en la búsqueda'}, status=500)
    
def profesores(request):
    profesores = Profesor.objects.all()

    profesores_concatenados = []
    for profesor in profesores:
        nombre_completo = f"{profesor.nombre} {profesor.segundo_nombre} {profesor.apellido} {profesor.segundo_apellido}"
        profesores_concatenados.append({
            'id': profesor.id_profesor,
            'nombre_completo': nombre_completo
        })

    return JsonResponse(profesores_concatenados, safe=False)
    
def modulo_por_id(request, modulo):
    try:
        modulo_obj = Modulo.objects.get(id_modulo=modulo)
        data = {
            'id_modulo': modulo_obj.id_modulo,
            'hora_modulo': modulo_obj.hora_modulo
        }
        return JsonResponse(data, status=200)
    except Modulo.DoesNotExist:
        logger.error(f'Módulo con ID {modulo} no encontrado.')
        return JsonResponse({'mensaje': 'Módulo no encontrado'}, status=404)
    except Exception as e:
        logger.error(f'Error al realizar la búsqueda: {str(e)}')
        return JsonResponse({'mensaje': 'Error en la búsqueda'}, status=500)
    
def actualizar_reemplazo(request):  
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            reemplazo_id = data.get('reemplazoId')
            semana = data.get('semana')
            profesor_remp = data.get('profesor_remp')
            
            if not reemplazo_id or not semana or not profesor_remp:
                return JsonResponse({'success': False, 'message': 'Faltan datos requeridos.'}, status=400)
            try:
                reemplazo = Reemplazos.objects.get(id_reemplazo=reemplazo_id)
                logger.debug("Reemplazo encontrado: %s", reemplazo)
            except Reemplazos.DoesNotExist:
                logger.error("Reemplazo no encontrado con id_reemplazo=%s", reemplazo_id)
                return JsonResponse({'success': False, 'message': 'Reemplazo no encontrado.'}, status=404)
            reemplazo.semana = semana
            reemplazo.profesor_reemplazo = profesor_remp
            reemplazo.save()
            return JsonResponse({'success': True, 'message': 'Reemplazo actualizado exitosamente.'})

        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'message': 'Error al procesar el JSON.'}, status=400)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)
    return JsonResponse({'success': False, 'message': 'Método no permitido.'}, status=405)

def obtenerHorario(request,id_profesor):
    horarios = Horario.objects.filter(profesor_id_profesor=id_profesor, activo=1)
    horarios_data = []
    for horario in horarios:
        horarios_data.append({
            'id_horario': horario.id_horario,
            'seccion': horario.seccion,
            'jornada': horario.jornada,
            'asignatura': horario.asignatura_id_asignatura.nombre_asignatura,
            'id_asignatura': horario.asignatura_id_asignatura.id_asignatura,
            'sala': horario.sala_id_sala.numero_sala,
            'id_sala': horario.sala_id_sala.id_sala,
            'dia': horario.dia_semana_id_dia.nombre_dia,
            'id_dia': horario.dia_semana_id_dia.id_dia,
            'modulo': horario.modulo_id_modulo.hora_modulo,
            'id_modulo': horario.modulo_id_modulo.id_modulo,
            'id_carrera': horario.carrera_id_carrera.id_carrera
        })
    return JsonResponse({'horarios': horarios_data})

def horas_periodo(request):
    docente_id = request.GET.get('profesorId')
    fecha_inicio = request.GET.get('fechaInicio')
    fecha_termino = request.GET.get('fechaTermino')

    if not docente_id or not fecha_inicio or not fecha_termino:
        logger.error("Faltan parámetros en la solicitud.")
        return JsonResponse({'error': 'Faltan parámetros'}, status=400)

    try:
        fecha_inicio = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
        fecha_termino = datetime.strptime(fecha_termino, '%Y-%m-%d').date()

        if fecha_termino < fecha_inicio:
            logger.warning("La fecha de término no puede ser antes de la fecha de inicio.")
            return JsonResponse({'error': 'La fecha de término no puede ser antes de la fecha de inicio'}, status=400)

        try:
            profesor = Profesor.objects.get(id_profesor=docente_id)
            nombre_completo = f"{profesor.nombre} {profesor.segundo_nombre} {profesor.apellido} {profesor.segundo_apellido}".strip()
            logger.debug(f"Profesor encontrado: {nombre_completo}")
        except Profesor.DoesNotExist:
            logger.error(f"Profesor con id {docente_id} no encontrado.")
            return JsonResponse({'error': 'Profesor no encontrado'}, status=404)

        dias_de_rango = [dia.id_dia for dia in DiaSemana.objects.filter(id_dia__in=range(1, 8))]

        horarios = Horario.objects.filter(
            profesor_id_profesor=profesor,
            dia_semana_id_dia__in=dias_de_rango
        )

        horas_horarios = horarios.count()

        with connection.cursor() as cursor:
            cursor.execute('''
    select
    concat(p.nombre,' ',IF(p.segundo_nombre IS NOT NULL, CONCAT(p.segundo_nombre, ' '),''), p.apellido, IF(p.segundo_apellido IS NOT NULL, CONCAT(' ',p.segundo_apellido),'')) as profesor,
    a.nombre_asignatura,
    c.nombre_carrera,
    h.jornada,
    h.seccion,
    s.numero_sala,
        CONCAT(
            SUBSTRING_INDEX(GROUP_CONCAT(mo.hora_modulo ORDER BY mo.hora_modulo SEPARATOR ', '), '-', 1), 
            '-',
            SUBSTRING_INDEX(SUBSTRING_INDEX(GROUP_CONCAT(mo.hora_modulo ORDER BY mo.hora_modulo SEPARATOR ', '), ',', -1), '-', -1)
        ) AS modulo,
    ds.nombre_dia,
    r.fecha_reemplazo, 
    r.profesor_reemplazo,
    l.motivo,
    l.observaciones,
    GROUP_CONCAT(r.id_reemplazo ORDER BY r.id_reemplazo SEPARATOR '-') as id_reemplazos
    from test.reemplazos r
    join test.horario h on h.id_horario = r.horario_id_horario 
    join test.profesor p on p.id_profesor = h.profesor_id_profesor
    join test.sala s on s.id_sala = h.sala_id_sala 
    join test.modulo mo on mo.id_modulo = h.modulo_id_modulo
    join test.asignatura a on a.id_asignatura = h.asignatura_id_asignatura 
    join test.dia_semana ds on ds.id_dia = h.dia_semana_id_dia
    join test.carrera c on h.carrera_id_carrera = c.id_carrera
    join test.licencia l on l.profesor_id_profesor = p.id_profesor
    where r.fecha_reemplazo between l.fecha_inicio and l.fecha_termino and h.activo = true and l.estado = "asignado" 
    group by r.semana, r.fecha_reemplazo, r.profesor_reemplazo, s.numero_sala, h.seccion, h.profesor_id_profesor, h.dia_semana_id_dia
    order by r.fecha_reemplazo
            ''', [fecha_inicio, fecha_termino, nombre_completo])
            reemplazos = cursor.fetchall()

        reemplazos_listados = []

        for row in reemplazos:
            reemplazos_listados.append({
                'semana': row[0],
                'profesor_nombre': row[1],
                'nombre_asignatura': row[2],
                'seccion': row[3],
                'numero_sala': row[4],
                'hora_modulo': row[5],
                'registros': row[6],
                'nombre_dia': row[7],
                'fecha_reemplazo': row[8],
                'profesor_reemplazo': row[9],
                'id_reemplazo': row[10],
                'id_modulo': row[11],
                'id_dia': row[12]
            })

        horas_reemplazo = sum(reemplazo['registros'] for reemplazo in reemplazos_listados)
        horas_totales = horas_horarios + horas_reemplazo

        tipo_pago = "BONO" if horas_totales >= 40 else "PROGRAMACIÓN"

        return JsonResponse({
            'horas_horarios': horas_horarios,
            'horas_reemplazo': horas_reemplazo,
            'horas_totales': horas_totales,
            'tipo_pago': tipo_pago,
            'reemplazos': reemplazos_listados
        })

    except Exception as e:
        logger.error(f"Error inesperado: {str(e)}")
        return JsonResponse({'error': f'Error: {str(e)}'}, status=500)

def copy_cell_border(from_cell, to_cell):
    to_cell.border = Border(
        left=from_cell.border.left,
        right=from_cell.border.right,
        top=from_cell.border.top,
        bottom=from_cell.border.bottom,
        diagonal=from_cell.border.diagonal,
        diagonal_direction=from_cell.border.diagonal_direction,
        outline=from_cell.border.outline,
        vertical=from_cell.border.vertical,
        horizontal=from_cell.border.horizontal
    )

def reporte_dara(request):
    licenciaID = request.GET.get('licenciaID')
    fechaInicio = request.GET.get('fechaInicio')
    fechaTermino = request.GET.get('fechaTermino')

    sql_query = '''
    SELECT
        CONCAT(p.nombre, ' ', IF(p.segundo_nombre IS NOT NULL, CONCAT(p.segundo_nombre, ' '), ''), p.apellido, IF(p.segundo_apellido IS NOT NULL, CONCAT(' ', p.segundo_apellido), '')) AS profesor,
        a.nombre_asignatura,
        c.nombre_carrera,
        h.jornada,
        h.seccion,
        s.numero_sala,
        CONCAT(
            SUBSTRING_INDEX(GROUP_CONCAT(mo.hora_modulo ORDER BY mo.hora_modulo SEPARATOR ', '), '-', 1),
            '-',
            SUBSTRING_INDEX(SUBSTRING_INDEX(GROUP_CONCAT(mo.hora_modulo ORDER BY mo.hora_modulo SEPARATOR ', '), ',', -1), '-', -1)
        ) AS modulo,
        ds.nombre_dia,
    CONCAT(MIN(r.fecha_reemplazo), '/', MAX(r.fecha_reemplazo)) AS fechas_reemplazo,
        r.profesor_reemplazo,
        l.motivo,
        l.observaciones
    FROM test.reemplazos r
    JOIN test.horario h ON h.id_horario = r.horario_id_horario
    JOIN test.profesor p ON p.id_profesor = h.profesor_id_profesor
    JOIN test.sala s ON s.id_sala = h.sala_id_sala
    JOIN test.modulo mo ON mo.id_modulo = h.modulo_id_modulo
    JOIN test.asignatura a ON a.id_asignatura = h.asignatura_id_asignatura
    JOIN test.dia_semana ds ON ds.id_dia = h.dia_semana_id_dia
    JOIN test.carrera c ON h.carrera_id_carrera = c.id_carrera
    JOIN test.licencia l ON l.profesor_id_profesor = p.id_profesor
    WHERE r.fecha_reemplazo BETWEEN %s AND %s
    AND l.id_licencia = %s
    AND h.activo = TRUE
    AND l.estado = "asignado"
    GROUP BY c.nombre_carrera, h.jornada,h.seccion,s.numero_sala,r.profesor_reemplazo,a.nombre_asignatura,ds.nombre_dia,l.motivo,l.observaciones
    ORDER BY r.fecha_reemplazo, mo.hora_modulo
    '''
    
    with connection.cursor() as cursor:
        cursor.execute(sql_query, [fechaInicio, fechaTermino, licenciaID])
        rows = cursor.fetchall()

    reemplazos_listados = []
    for row in rows:
        modulo = row[6].split('-')
        inicio_modulo = modulo[0]
        fin_modulo = modulo[1]
        modulo_fec = row[8].split('/')
        fecha_inicio = modulo_fec[0]
        fecha_termino = modulo_fec[1]
        
        reemplazos_listados.append({
            'profesor': row[0],
            'nombre_asignatura': row[1],
            'nombre_carrera': row[2],
            'jornada': row[3],
            'seccion': row[4],
            'numero_sala': row[5],
            'inicio_modulo': inicio_modulo,
            'fin_modulo': fin_modulo,
            'nombre_dia': row[7],
            'fecha_reemplazo': row[8],
            'fecha_inicio': fecha_inicio,
            'fecha_termino': fecha_termino,
            'profesor_reemplazo': row[9],
            'motivo': row[10],
            'observaciones': row[11]
        })

    asignaturas_y_carreras = {}
    for reemplazo in reemplazos_listados:
        asignatura = reemplazo['nombre_asignatura']
        carrera = reemplazo['nombre_carrera']
        key = (asignatura, carrera)

        if key not in asignaturas_y_carreras:
            asignaturas_y_carreras[key] = []

        asignaturas_y_carreras[key].append(reemplazo)

    base_dir = os.path.dirname(os.path.abspath(__file__))
    file_paths = []

    for (asignatura, carrera), reemplazos in asignaturas_y_carreras.items():
        file_path = os.path.join(base_dir, 'static', 'documents', 'DARA.xlsx')
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        row_actual = 34
        row_reemplazo = 46
        
        for reemplazo in reemplazos:
            ws.merge_cells(f'N{row_actual}:Q{row_actual}')
            ws.merge_cells(f'N{row_reemplazo}:Q{row_reemplazo}')
            ws.merge_cells(f'J{row_actual}:K{row_actual}')
            ws.merge_cells(f'J{row_reemplazo}:K{row_reemplazo}')
            ws.merge_cells(f'H{row_actual}:I{row_actual}')
            ws.merge_cells(f'H{row_reemplazo}:I{row_reemplazo}')
            ws[f'E10'] = reemplazo['nombre_asignatura']
            ws[f'E14'] = reemplazo['seccion']
            ws[f'K14'] = reemplazo['jornada']
            ws[f'E18'] = reemplazo['nombre_carrera']
            ws[f'E28'] = reemplazo['observaciones']

            ws[f'C{row_actual}'] = reemplazo['inicio_modulo']
            ws[f'D{row_actual}'] = reemplazo['fin_modulo']
            ws[f'E{row_actual}'] = reemplazo['nombre_dia']
            ws[f'F{row_actual}'] = reemplazo['numero_sala']
            ws[f'H{row_actual}'] = reemplazo['fecha_inicio']
            ws[f'J{row_actual}'] = reemplazo['fecha_termino']
            ws[f'N{row_actual}'] = reemplazo['profesor']

            ws[f'C{row_reemplazo}'] = reemplazo['inicio_modulo']
            ws[f'D{row_reemplazo}'] = reemplazo['fin_modulo']
            ws[f'E{row_reemplazo}'] = reemplazo['nombre_dia']
            ws[f'F{row_reemplazo}'] = reemplazo['numero_sala']
            ws[f'H{row_reemplazo}'] = reemplazo['fecha_inicio']
            ws[f'J{row_reemplazo}'] = reemplazo['fecha_termino']
            ws[f'N{row_reemplazo}'] = reemplazo['profesor_reemplazo']

            row_actual += 1
            row_reemplazo += 1
            if row_actual > 34+8:
                ws.insert_rows(row_actual+1)
                ws.insert_rows(row_reemplazo+1)

        file_name = f"{asignatura}_{carrera}.xlsx"
        output_file_path = os.path.join(base_dir, 'static', 'documents', file_name)

        wb.save(output_file_path)
        file_paths.append(output_file_path)

    return JsonResponse({"files": file_paths, "reemplazos_listados": reemplazos_listados}, safe=False)